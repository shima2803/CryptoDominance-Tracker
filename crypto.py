# crypto_top10_xlsx_dominance.py
# ============================================================
# Top 10 Criptomoedas (USD) - CoinGecko
# - Calcula dominância (%) sobre o market cap global
# - Imprime no terminal
# - Salva Excel (.xlsx) na Área de Trabalho
# - Substitui o arquivo se já existir
#
# Dependências:
#   pip install requests openpyxl
# ============================================================

import os
import sys
import time
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

COINGECKO_MARKETS_URL = "https://api.coingecko.com/api/v3/coins/markets"
COINGECKO_GLOBAL_URL = "https://api.coingecko.com/api/v3/global"

HEADERS = {"User-Agent": "crypto-dominance-script/1.0"}
XLSX_NAME = "top10_cripto_usd.xlsx"


# ============================================================
# Utils
# ============================================================
def get_desktop_path() -> str:
    home = os.path.expanduser("~")
    desktop = os.path.join(home, "Desktop")
    return desktop if os.path.isdir(desktop) else home


# ============================================================
# API
# ============================================================
def fetch_global_market_cap():
    r = requests.get(COINGECKO_GLOBAL_URL, headers=HEADERS, timeout=30)
    r.raise_for_status()
    data = r.json()
    return data["data"]["total_market_cap"]["usd"]


def fetch_top10_crypto(retries=3, backoff=2):
    params = {
        "vs_currency": "usd",
        "order": "market_cap_desc",
        "per_page": 10,
        "page": 1,
        "sparkline": "false",
        "price_change_percentage": "24h",
    }

    last_err = None
    for attempt in range(1, retries + 1):
        try:
            r = requests.get(COINGECKO_MARKETS_URL, headers=HEADERS, params=params, timeout=30)

            if r.status_code == 429:
                time.sleep(backoff * attempt)
                continue

            r.raise_for_status()
            data = r.json()

            if not isinstance(data, list) or not data:
                raise ValueError("Resposta inválida da API.")

            return [{
                "rank": c.get("market_cap_rank"),
                "name": c.get("name"),
                "symbol": (c.get("symbol") or "").upper(),
                "price": c.get("current_price"),
                "market_cap": c.get("market_cap"),
                "change_24h": c.get("price_change_percentage_24h"),
            } for c in data[:10]]

        except Exception as e:
            last_err = e
            if attempt < retries:
                time.sleep(backoff * attempt)

    raise RuntimeError(f"Falha ao buscar dados: {last_err}")


# ============================================================
# Terminal
# ============================================================
def print_terminal(data, global_cap):
    print("\n==============================")
    print("TOP 10 CRIPTOMOEDAS (USD)")
    print("Dominancia sobre o mercado global")
    print("==============================\n")

    for c in data:
        dom = (c["market_cap"] / global_cap * 100) if c["market_cap"] else 0
        price = c["price"]
        price_str = f"${price:,.6f}" if price < 1 else f"${price:,.2f}"
        chg = c["change_24h"]
        chg_str = f"{chg:.2f}%" if isinstance(chg, (int, float)) else "-"

        print(
            f"{c['rank']:>2} | {c['name']:<18} ({c['symbol']:<6}) | "
            f"Preco: {price_str:<14} | Dom: {dom:>5.2f}% | 24h: {chg_str}"
        )


# ============================================================
# Excel
# ============================================================
def save_xlsx(data, global_cap, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Top10_Cripto_USD"

    ws.append([
        "DataColeta",
        "Rank",
        "Cripto",
        "Simbolo",
        "PrecoUSD",
        "MarketCapUSD",
        "DominanciaPct",
        "Variacao24h"
    ])

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for c in data:
        dominance = (c["market_cap"] / global_cap * 100) if c["market_cap"] else 0
        ws.append([
            now,
            c["rank"],
            c["name"],
            c["symbol"],
            c["price"],
            c["market_cap"],
            dominance,
            c["change_24h"],
        ])

    ws.freeze_panes = "A2"

    for col in range(1, ws.max_column + 1):
        max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 30)

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=5).number_format = "$#,##0.00"
        ws.cell(row=r, column=6).number_format = "$#,##0"
        ws.cell(row=r, column=7).number_format = '0.00"%"'
        ws.cell(row=r, column=8).number_format = '0.00"%"'

    if os.path.exists(out_path):
        try:
            os.remove(out_path)
        except PermissionError:
            raise PermissionError("Feche o Excel antes de atualizar o arquivo.")

    wb.save(out_path)


# ============================================================
# Main
# ============================================================
def main():
    try:
        global_cap = fetch_global_market_cap()
        data = fetch_top10_crypto()

        print_terminal(data, global_cap)

        desktop = get_desktop_path()
        out_file = os.path.join(desktop, XLSX_NAME)

        save_xlsx(data, global_cap, out_file)

        print(f"\nArquivo Excel atualizado em:\n{out_file}\n")

    except Exception as e:
        print(f"\nERRO: {e}\n")
        sys.exit(1)


if __name__ == "__main__":
    main()
